# Importações
import sqlite3
import openpyxl

# Conexão com banco de dados
bd = sqlite3.connect("database.db")
cursor = bd.cursor()
 
res = str(input("digite algo ")) 
# Scripts dos banco de dados
Admin = """
create view adm as
select * from ( 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202201' a 
left join 'inf_mensal_fidc_tab_IV_202201' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202202' a 
left join 'inf_mensal_fidc_tab_IV_202202' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202203' a 
left join 'inf_mensal_fidc_tab_IV_202203' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202204' a 
left join 'inf_mensal_fidc_tab_IV_202204' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202205' a 
left join 'inf_mensal_fidc_tab_IV_202205' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202206' a 
left join 'inf_mensal_fidc_tab_IV_202206' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202207' a 
left join 'inf_mensal_fidc_tab_IV_202207' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202208' a 
left join 'inf_mensal_fidc_tab_IV_202208' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202209' a 
left join 'inf_mensal_fidc_tab_IV_202209' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202210' a 
left join 'inf_mensal_fidc_tab_IV_202210' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202211' a 
left join 'inf_mensal_fidc_tab_IV_202211' b
on a.cnpj_fundo = b.cnpj_fundo
union all 

select a.*, b.tab_iv_a_vl_pl, b.tab_iv_b_vl_pl_medio 
from 'inf_mensal_fidc_tab_I_202212' a 
left join 'inf_mensal_fidc_tab_IV_202212' b
on a.cnpj_fundo = b.cnpj_fundo

)AS derived_table
WHERE TAB_IV_A_VL_PL <= 2000000000 AND TAB_IV_A_VL_PL >= 1000000
ORDER BY TAB_IV_A_VL_PL DESC;"""

User = """
create view user as
select * from ( select * 
from 'inf_mensal_fidc_tab_IV_202112' 
union all select *
from 'inf_mensal_fidc_tab_IV_202201' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202202' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202203' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202204' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202205' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202206' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202207' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202208' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202209' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202210' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202211' 
union all select * 
from 'inf_mensal_fidc_tab_IV_202212'
) AS derived_table
where TAB_IV_A_VL_PL <= 2000000000 and TAB_IV_A_VL_PL >= 1000000
order by TAB_IV_A_VL_PL desc"""

# Teste para visualização de tabelas ou dados 
if res == "a":
    a=0
    CNPJ = 0
    NOME = 1
    DATA = 2
    PL_MES = 3
    PL_MEDIO = 4
    cursor.execute(Admin)
    resultados = cursor.fetchall()
    # "description" pega os nomes das tabelas
    column_names = [description[0] for description in cursor.description]
    for row in resultados:
        if resultados[a][DATA][0:4] == "2021":
            dez_21 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "01":
            jan_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "02":
            fev_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "03":
            mar_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "04":
            abr_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "05":
            mai_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "06":
            jun_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "07":
            jul_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "08":
            ago_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "09":
            set_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "10":
            out_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "11":
            nov_22 = resultados[a][DATA][:7]
        if resultados[a][DATA][5:7] == "12":
            dez_22 = resultados[a][DATA][:7]
        print(column_names[a])
        a+=1
    
    bd.commit()
    bd.close()
       
if res == "b":
    try:
        a=0
        cursor.execute(User)
        resultados = cursor.fetchall()
        column_names = [description[0] for description in cursor.description]

        # Criando uma planilha e selecionando uma tabela ativa
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for i, name in enumerate(column_names):
            worksheet.cell(row=1, column=i+1, value=name)

        for row_num, row_data in enumerate(resultados, start=2):
            for col_num, cell_data in enumerate(row_data):
                worksheet.cell(row=row_num, column=col_num+1, value=cell_data)

        # Salvando a planilha (precisa alterar o nome da admin)
        workbook.save('User.xlsx')

    except Exception as e:
        print(f"Error: {e}")
    bd.commit()
    bd.close()
    
